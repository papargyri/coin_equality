"""
Utilities for comparing multiple optimization runs.

Provides functions for:
- Discovering result directories from path patterns
- Loading and validating CSV files
- Generating case names from directory paths
- Creating Excel comparison workbooks
"""

import glob
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def discover_result_directories(path_patterns):
    """
    Discover valid result directories from path patterns.

    Expands glob patterns and validates that each directory contains
    the required optimization_summary.csv file.

    Parameters
    ----------
    path_patterns : list of str
        List of directory paths or glob patterns
        Examples: ['results/baseline/', 'results/test_*/']

    Returns
    -------
    list of Path
        Sorted list of directories containing optimization_summary.csv

    Raises
    ------
    ValueError
        If no valid directories found

    Examples
    --------
    >>> dirs = discover_result_directories(['results/test_*_000/'])
    >>> dirs = discover_result_directories(['results/baseline/', 'results/high_eta/'])
    """
    directories = []

    for pattern in path_patterns:
        matches = glob.glob(pattern)
        for match in matches:
            path = Path(match)
            if path.is_dir() and (path / 'optimization_summary.csv').exists():
                directories.append(path)

    if not directories:
        raise ValueError(f"No valid result directories found for patterns: {path_patterns}")

    return sorted(set(directories))


def generate_case_name(directory_path):
    """
    Generate readable case name from directory path.

    Uses the directory name as the case name for display in plots and tables.

    Parameters
    ----------
    directory_path : Path
        Directory path

    Returns
    -------
    str
        Case name for display

    Examples
    --------
    >>> generate_case_name(Path('results/baseline'))
    'baseline'
    >>> generate_case_name(Path('results/test_010'))
    'test_010'
    """
    return directory_path.name


def load_optimization_summaries(directories):
    """
    Load optimization_summary.csv from multiple directories.

    Parameters
    ----------
    directories : list of Path
        Result directories

    Returns
    -------
    dict
        {case_name: pd.DataFrame, ...}
        Each DataFrame contains optimization summary with columns:
        iteration, n_evaluations, objective, termination_status, ...
        (elapsed_time included if available)

    Examples
    --------
    >>> dirs = [Path('results/baseline'), Path('results/test_010')]
    >>> data = load_optimization_summaries(dirs)
    >>> data.keys()
    dict_keys(['baseline', 'test_010'])
    """
    data = {}
    for directory in directories:
        case_name = generate_case_name(directory)
        csv_path = directory / 'optimization_summary.csv'

        # Parse custom CSV format - find iteration history section
        with open(csv_path, 'r') as f:
            lines = f.readlines()

        # Find the "Iterative Refinement - Iteration History" section
        start_idx = None
        for i, line in enumerate(lines):
            if 'Iterative Refinement - Iteration History' in line:
                start_idx = i + 1  # Skip section header, header is next line
                break

        if start_idx is None:
            raise ValueError(f"Could not find iteration history in {csv_path}")

        # Read header and data rows
        header_line = lines[start_idx].strip()
        data_start = start_idx + 1

        # Read rows until blank line or next section
        data_rows = []
        for i in range(data_start, len(lines)):
            line = lines[i].strip()
            if not line or line.startswith('Iterative'):
                break
            data_rows.append(line)

        # Parse into DataFrame
        from io import StringIO
        csv_content = header_line + '\n' + '\n'.join(data_rows)
        df = pd.read_csv(StringIO(csv_content))

        # Rename columns to match expected format
        column_mapping = {
            'Iteration': 'iteration',
            'Objective': 'objective',
            'Evaluations': 'n_evaluations',
            'Status': 'termination_status'
        }
        df = df.rename(columns=column_mapping)

        data[case_name] = df

    return data


def load_control_points(directories):
    """
    Load control points for each iteration from optimization_summary.csv files.

    Parameters
    ----------
    directories : list of Path
        Result directories

    Returns
    -------
    tuple
        (f_data, s_data) where each is:
        {case_name: {iteration: pd.DataFrame(columns=['t', 'f' or 's']), ...}, ...}
        Nested dicts with iteration-specific control point DataFrames for each case.

    Examples
    --------
    >>> dirs = [Path('results/baseline'), Path('results/test_010')]
    >>> f_data, s_data = load_control_points(dirs)
    >>> f_data['baseline'][1]  # f control points for iteration 1
       t      f
    0  0.0  0.50
    1  400.0  0.50
    """
    f_data = {}
    s_data = {}

    for directory in directories:
        case_name = generate_case_name(directory)
        csv_path = directory / 'optimization_summary.csv'

        with open(csv_path, 'r') as f:
            lines = f.readlines()

        case_f_iterations = {}
        case_s_iterations = {}

        # Find all "Iteration N Control Points (f):" and "Control Points (s):" sections
        for i, line in enumerate(lines):
            is_f = 'Control Points (f):' in line
            is_s = 'Control Points (s):' in line

            if is_f or is_s:
                # Extract iteration number from line like "Iteration 1 Control Points (f):"
                parts = line.split()
                if 'Iteration' in parts:
                    iter_idx = parts.index('Iteration') + 1
                    if iter_idx < len(parts):
                        iteration_num = int(parts[iter_idx])
                    else:
                        continue
                else:
                    continue

                # Determine variable name to parse
                var_name = 'f' if is_f else 's'

                # Parse control points - format: "  t=XXX.X yr: f=Y.YYYYYY" or "  t=XXX.X yr: s=Y.YYYYYY"
                control_points = []
                j = i + 1
                while j < len(lines):
                    cp_line = lines[j].strip()
                    if not cp_line or 'Iteration' in cp_line or 'Control Points' in cp_line:
                        break
                    # Parse "t=XXX.X yr: {var}=Y.YYYYYY"
                    if 't=' in cp_line and f'{var_name}=' in cp_line:
                        try:
                            t_part = cp_line.split('t=')[1].split('yr:')[0].strip()
                            val_part = cp_line.split(f'{var_name}=')[1].strip()
                            t_val = float(t_part)
                            val = float(val_part)
                            control_points.append({'t': t_val, var_name: val})
                        except (ValueError, IndexError):
                            pass
                    j += 1

                if control_points:
                    df = pd.DataFrame(control_points)
                    if is_f:
                        case_f_iterations[iteration_num] = df
                    else:
                        case_s_iterations[iteration_num] = df

        f_data[case_name] = case_f_iterations
        s_data[case_name] = case_s_iterations

    return f_data, s_data


def load_results_csvs(directories):
    """
    Load results.csv from multiple directories.

    Only includes cases where results.csv exists. Prints warning for
    directories missing results.csv.

    Parameters
    ----------
    directories : list of Path
        Result directories

    Returns
    -------
    dict
        {case_name: pd.DataFrame, ...}
        Only includes cases where results.csv exists.
        Each DataFrame contains model results with columns:
        t, y, K, T_atm, E_cum, etc.

    Examples
    --------
    >>> dirs = [Path('results/baseline'), Path('results/test_010')]
    >>> data = load_results_csvs(dirs)
    """
    data = {}
    for directory in directories:
        case_name = generate_case_name(directory)
        csv_path = directory / 'results.csv'
        if csv_path.exists():
            data[case_name] = pd.read_csv(csv_path)
        else:
            print(f"Warning: results.csv not found in {directory}, skipping results comparison for this case")

    return data


def create_directories_sheet(wb, directories):
    """
    Create sheet listing all directories included in comparison.

    Parameters
    ----------
    wb : openpyxl.Workbook
        Workbook to add sheet to
    directories : list of Path
        List of result directories
    """
    ws = wb.create_sheet('Directories')

    # Header row
    ws['A1'] = 'Case Name'
    ws['A1'].font = Font(bold=True)
    ws['A1'].fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

    ws['B1'] = 'Directory Path'
    ws['B1'].font = Font(bold=True)
    ws['B1'].fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

    # Data rows
    for row_idx, directory in enumerate(directories, start=2):
        case_name = generate_case_name(directory)
        ws.cell(row_idx, 1, case_name)
        ws.cell(row_idx, 2, str(directory))


def create_control_point_sheets(wb, control_points_data, var_name):
    """
    Create sheets showing control points for each iteration across all cases.

    Parameters
    ----------
    wb : openpyxl.Workbook
        Workbook to add sheets to
    control_points_data : dict
        {case_name: {iteration: DataFrame(columns=['t', var_name]), ...}, ...}
    var_name : str
        Variable name ('f' or 's')
    """
    if not control_points_data:
        return

    # Find all unique iterations across all cases
    all_iterations = set()
    for case_iters in control_points_data.values():
        all_iterations.update(case_iters.keys())

    if not all_iterations:
        return

    # Create a sheet for each iteration
    for iteration in sorted(all_iterations):
        sheet_name = f'Iter {iteration} {var_name}(t)'
        ws = wb.create_sheet(sheet_name)

        # Header row
        ws['A1'] = 'Time (years)'
        ws['A1'].font = Font(bold=True)
        ws['A1'].fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

        # Add case names as column headers
        case_names = []
        for col_idx, case_name in enumerate(sorted(control_points_data.keys()), start=2):
            if iteration in control_points_data[case_name]:
                case_names.append((col_idx, case_name))
                cell = ws.cell(1, col_idx, case_name)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')

        # Collect all unique time points for this iteration
        time_points = set()
        for case_name, case_iters in control_points_data.items():
            if iteration in case_iters:
                df = case_iters[iteration]
                time_points.update(df['t'].values)

        if not time_points:
            continue

        # Sort time points
        sorted_times = sorted(time_points)

        # Write data rows
        for row_idx, t_val in enumerate(sorted_times, start=2):
            ws.cell(row_idx, 1, t_val)

            # For each case, write value at this time point (if exists)
            for col_idx, case_name in case_names:
                if iteration in control_points_data[case_name]:
                    df = control_points_data[case_name][iteration]
                    # Find value at this time
                    matching_rows = df[df['t'] == t_val]
                    if not matching_rows.empty:
                        val = matching_rows.iloc[0][var_name]
                        cell = ws.cell(row_idx, col_idx, val)
                        cell.number_format = '0.000000'  # 6 decimal places


def create_comparison_xlsx(optimization_data, directories, output_path):
    """
    Create Excel workbook comparing optimization summaries across cases.

    Creates multi-sheet workbook with one sheet per metric (Objective,
    Evaluations, Elapsed Time, Termination Status) plus control point sheets
    for each iteration. Cases are columns, iterations are rows.

    Parameters
    ----------
    optimization_data : dict
        {case_name: optimization_summary_df, ...}
    directories : list of Path
        List of result directories included in comparison
    output_path : Path or str
        Output Excel file path

    Notes
    -----
    Workbook structure:
    - Sheet 1: "Directories" - list of all compared directories
    - Sheet 2: "Objective" - objective values by iteration
    - Sheet 3: "Evaluations" - function evaluation counts
    - Sheet 4: "Elapsed Time (s)" - computation time (if available)
    - Sheet 5: "Termination Status" - optimization termination reasons
    - Sheets 6+: "Iter N f(t)" - f control points for each iteration
    - Additional sheets: "Iter N s(t)" - s control points (if dual optimization)

    Each metric sheet has:
    - Column A: Iteration number
    - Columns B+: One column per case
    - Header row with case names

    Each control point sheet has:
    - Column A: Time (years)
    - Columns B+: f or s values for each case at each time point
    - Header row with case names

    Examples
    --------
    >>> create_comparison_xlsx(opt_data, dirs, 'comparison_summary.xlsx')
    Comparison Excel workbook saved to: comparison_summary.xlsx
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Sheet 1: List of directories
    create_directories_sheet(wb, directories)

    # Load control points data (f and s)
    f_control_points, s_control_points = load_control_points(directories)

    # Define comparison sheets with column name and Excel number format
    # Excel number format codes (not Python format strings)
    comparison_specs = [
        ('Objective', 'objective', '0.000000E+00'),       # Scientific notation, 6 decimals
        ('Evaluations', 'n_evaluations', '0'),            # Integer
        ('Termination Status', 'termination_status', None)  # Text, no format
    ]

    # Add elapsed_time sheet only if data contains it
    if optimization_data and any('elapsed_time' in df.columns for df in optimization_data.values()):
        comparison_specs.insert(2, ('Elapsed Time (s)', 'elapsed_time', '0.00'))  # 2 decimal places

    for sheet_name, column_name, number_format in comparison_specs:
        create_comparison_sheet(wb, sheet_name, column_name, optimization_data, number_format)

    # Add control point sheets for f (abatement allocation)
    create_control_point_sheets(wb, f_control_points, 'f')

    # Add control point sheets for s (savings rate) if any case has s optimization
    if any(case_s_iters for case_s_iters in s_control_points.values()):
        create_control_point_sheets(wb, s_control_points, 's')

    # Auto-size columns
    for sheet in wb.worksheets:
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[column_letter].width = max_length + 2

    wb.save(output_path)
    wb.close()
    print(f"Comparison Excel workbook saved to: {output_path}")


def create_comparison_sheet(wb, sheet_name, column_name, data, number_format):
    """
    Create a single comparison sheet in the workbook.

    Parameters
    ----------
    wb : openpyxl.Workbook
        Workbook to add sheet to
    sheet_name : str
        Name of the sheet
    column_name : str
        Column name from optimization_summary.csv to extract
    data : dict
        {case_name: optimization_summary_df, ...}
    number_format : str or None
        Excel number format code (e.g., '0.00E+00', '0.00') or None for no formatting
    """
    ws = wb.create_sheet(sheet_name)

    # Header row with styling
    ws['A1'] = 'Iteration'
    ws['A1'].font = Font(bold=True)
    ws['A1'].fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

    case_names = list(data.keys())
    for col_idx, case_name in enumerate(case_names, start=2):
        cell = ws.cell(1, col_idx, case_name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    max_iterations = max(len(df) for df in data.values())

    for row_idx in range(max_iterations):
        # Iteration number in column A
        ws.cell(row_idx + 2, 1, row_idx + 1)

        # Values for each case in subsequent columns
        for col_idx, (case_name, df) in enumerate(data.items(), start=2):
            if row_idx < len(df) and column_name in df.columns:
                value = df.iloc[row_idx][column_name]
                cell = ws.cell(row_idx + 2, col_idx)

                if isinstance(value, (int, float)):
                    cell.value = value
                    # Apply Excel number format if provided
                    if number_format is not None:
                        cell.number_format = number_format
                else:
                    cell.value = str(value)


def clean_column_names(df):
    """
    Extract simple variable names from CSV headers with format 'var_name, Description, (units)'.

    Parameters
    ----------
    df : pd.DataFrame
        DataFrame with full column names from results.csv

    Returns
    -------
    pd.DataFrame
        DataFrame with simplified column names (just the variable name)
    """
    new_columns = []
    for col in df.columns:
        var_name = col.split(',')[0].strip()
        new_columns.append(var_name)
    df.columns = new_columns
    return df


def create_results_comparison_xlsx(results_data, directories, output_path):
    """
    Create Excel workbook comparing results time series across cases.

    Creates multi-sheet workbook with one sheet per variable. Each sheet has
    time in column A and one column per case.

    Parameters
    ----------
    results_data : dict
        {case_name: results_df, ...}
    directories : list of Path
        List of result directories included in comparison
    output_path : Path or str
        Output Excel file path

    Notes
    -----
    Workbook structure:
    - Sheet 1: "Directories" - list of all compared directories
    - Subsequent sheets: One sheet per variable (27 total)
      - Column A: Time (years)
      - Columns B+: One column per case with variable values

    Variable sheets match the plots in comparison_plots.pdf:
    Economic: y, y_eff, K, Consumption, Savings, s, Y_gross, Y_net
    Climate: delta_T, E, Ecum
    Abatement/Damage: f, mu, Lambda, AbateCost, Omega, Climate_Damage, marginal_abatement_cost
    Inequality/Utility: Gini, Gini_climate, G_eff, U, discounted_utility
    Exogenous: A, L, sigma, theta1
    """
    if not results_data:
        print("No results data available - skipping results comparison workbook")
        return

    results_data_cleaned = {name: clean_column_names(df.copy()) for name, df in results_data.items()}

    variable_specs = [
        ('y', 'Per-Capita Consumption'),
        ('y_eff', 'Effective Per-Capita Income'),
        ('K', 'Capital Stock'),
        ('Consumption', 'Total Consumption'),
        ('Savings', 'Gross Investment'),
        ('s', 'Savings Rate'),
        ('Y_gross', 'Gross Production'),
        ('Y_net', 'Net Output After Damages & Abatement'),
        ('delta_T', 'Temperature Change from Pre-Industrial'),
        ('E', 'CO2 Emissions Rate'),
        ('Ecum', 'Cumulative CO2 Emissions'),
        ('f', 'Abatement Allocation Fraction'),
        ('mu', 'Emissions Abatement Fraction'),
        ('Lambda', 'Abatement Cost (% of Output)'),
        ('AbateCost', 'Total Abatement Cost'),
        ('Omega', 'Climate Damage (% of Output)'),
        ('Climate_Damage', 'Total Climate Damage'),
        ('Gini', 'Starting Gini Index'),
        ('Gini_climate', 'Post-Climate-Damage Gini'),
        ('G_eff', 'Effective Gini Index'),
        ('U', 'Mean Utility Per Capita'),
        ('discounted_utility', 'Discounted Utility Per Capita'),
        ('marginal_abatement_cost', 'Marginal Abatement Cost'),
        ('A', 'Total Factor Productivity'),
        ('L', 'Population'),
        ('sigma', 'Carbon Intensity of GDP'),
        ('theta1', 'Abatement Cost Parameter (theta1)'),
    ]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    create_directories_sheet(wb, directories)

    for var_name, sheet_title in variable_specs:
        if not any(var_name in df.columns for df in results_data_cleaned.values()):
            continue

        ws = wb.create_sheet(sheet_title[:31])

        ws['A1'] = 'Time (years)'
        ws['A1'].font = Font(bold=True)
        ws['A1'].fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

        case_names = list(results_data_cleaned.keys())
        for col_idx, case_name in enumerate(case_names, start=2):
            cell = ws.cell(1, col_idx, case_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        representative_df = next(iter(results_data_cleaned.values()))
        time_values = representative_df['t'].values

        for row_idx, time_val in enumerate(time_values, start=2):
            ws.cell(row_idx, 1, float(time_val))

            for col_idx, (case_name, df) in enumerate(results_data_cleaned.items(), start=2):
                if var_name in df.columns:
                    value = df.iloc[row_idx - 2][var_name]
                    ws.cell(row_idx, col_idx, float(value))

    for sheet in wb.worksheets:
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[column_letter].width = max_length + 2

    wb.save(output_path)
    wb.close()
    print(f"Results comparison workbook saved to: {output_path}")
